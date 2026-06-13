"""
Report Generator Module for Salesforce Code Audit
Generates Excel reports with multiple sheets and optional markdown summary
"""

import json
import os
import re
from datetime import datetime
from typing import Dict, List, Any
from xml.sax.saxutils import escape
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from grading_engine import GradingResult, Grade


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TOOL_METADATA_PATH = os.path.join(SCRIPT_DIR, 'tool_version.json')
DEFAULT_TOOL_VERSION = '1.2.11'


def _load_tool_version() -> str:
    """Load the tool version from local metadata when available."""
    if os.path.exists(TOOL_METADATA_PATH):
        try:
            with open(TOOL_METADATA_PATH, 'r', encoding='utf-8') as handle:
                metadata = json.load(handle)
            version = str(metadata.get('version', '')).strip()
            if version:
                return version
        except (OSError, ValueError, TypeError):
            pass
    return DEFAULT_TOOL_VERSION


TOOL_VERSION = _load_tool_version()


class ExcelReportGenerator:
    """Generates formatted Excel reports for audit results"""
    
    # Comprehensive Rules Reference
    APEX_GURU_RULES = {
        'SOQL in Loop',
        'Schema.getGlobalDescribe() in Loop',
        'Schema.getGlobalDescribe() Not Efficient',
        'Redundant SOQL',
        'SObject Map in a For Loop',
        'SOQL with Negative Expressions',
        'SOQL with Unused Fields',
        'SOQL with Wildcard Filter',
        'SOQL Without a WHERE Clause or LIMIT Statement',
        'SOQL with Apex Filter',
        'Expensive Methods in Loop',
        'Expensive String Comparison',
        'Copying Elements with for Loop',
        'Sorting in Apex Instead of SOQL ORDER BY',
        'Busy Loop Delay',
        'Limits.getHeapSize() in Loop',
        'Unused Methods',
    }

    RULES_REFERENCE = [
        {
            'rule': 'SOQL in Loop',
            'category': 'Performance',
            'description': 'SOQL queries executed inside loops can hit governor limits (100 SOQL queries per transaction)',
            'severity': 'Critical',
            'recommendation': 'Move SOQL query outside the loop or use a SOQL-for loop to process records in bulk'
        },
        {
            'rule': 'DML in Loop',
            'category': 'Performance',
            'description': 'DML statements inside loops can hit governor limits (150 DML statements per transaction)',
            'severity': 'Critical',
            'recommendation': 'Collect records in a list and perform bulk DML after the loop'
        },
        {
            'rule': 'Indirect SOQL in Loop',
            'category': 'Performance',
            'description': 'Method calls inside loops that may contain SOQL queries',
            'severity': 'High',
            'recommendation': 'Refactor to avoid indirect SOQL execution in loops'
        },
        {
            'rule': 'Indirect DML in Loop',
            'category': 'Performance',
            'description': 'Method calls inside loops that may contain DML operations',
            'severity': 'High',
            'recommendation': 'Refactor to avoid indirect DML execution in loops'
        },
        {
            'rule': 'Non-Restrictive Query',
            'category': 'Performance',
            'description': 'SOQL queries without meaningful WHERE clauses can retrieve excessive records and hit heap limits',
            'severity': 'High',
            'recommendation': 'Add restrictive WHERE clauses and LIMIT clauses to reduce data volume'
        },
        {
            'rule': '@future Method Usage',
            'category': 'Architecture',
            'description': '@future methods are legacy async pattern, less flexible than Queueable',
            'severity': 'Medium',
            'recommendation': 'Use Queueable interface instead for better control and chaining'
        },
        {
            'rule': 'Async in Trigger',
            'category': 'Architecture',
            'description': '@future or Queueable called directly in triggers should be in handler classes',
            'severity': 'High',
            'recommendation': 'Move async logic to handler classes for better testability and maintainability'
        },
        {
            'rule': 'EventBus without Callback',
            'category': 'Best Practices',
            'description': 'Platform Events published without error handling callbacks can silently fail',
            'severity': 'High',
            'recommendation': 'Implement EventBus.TriggerContext callback for error handling'
        },
        {
            'rule': 'CMDT SOQL without Filter',
            'category': 'Performance',
            'description': 'Custom Metadata queries without filters retrieve all records unnecessarily',
            'severity': 'Medium',
            'recommendation': 'Use Custom_Metadata_Type__mdt.getInstance() or add WHERE clause'
        },
        {
            'rule': 'Hardcoded Salesforce ID',
            'category': 'Best Practices',
            'description': 'Hardcoded 15/18-char IDs break when migrating between orgs',
            'severity': 'High',
            'recommendation': 'Use Custom Metadata, Custom Settings, or dynamic SOQL to retrieve IDs'
        },
        {
            'rule': 'Missing CRUD/FLS Check',
            'category': 'Security',
            'description': 'Operations without CRUD/FLS checks can expose unauthorized data. Context-aware severity: Critical for public endpoints, Medium for internal services, Low for automation',
            'severity': 'Critical/Medium/Low',
            'recommendation': 'Use WITH USER_MODE (modern), WITH SECURITY_ENFORCED, Security.stripInaccessible(), or Schema checks'
        },
        {
            'rule': 'SOQL Injection Risk',
            'category': 'Security',
            'description': 'User input concatenated into SOQL queries can allow malicious queries',
            'severity': 'Critical',
            'recommendation': 'Use bind variables (:variable) or String.escapeSingleQuotes() for user input'
        },
        {
            'rule': 'Missing Sharing Keyword',
            'category': 'Security',
            'description': 'Classes without sharing keywords may bypass record-level security. Context-aware severity: Critical for public endpoints, Medium for internal services, Low for automation',
            'severity': 'Critical/Medium/Low',
            'recommendation': 'Add "with sharing" (enforce), "without sharing" (bypass), or "inherited sharing" (inherit from caller)'
        },
        {
            'rule': 'Hardcoded Credentials',
            'category': 'Security',
            'description': 'Hardcoded passwords, API keys, or tokens in code are a major security risk',
            'severity': 'Critical',
            'recommendation': 'Use Named Credentials, Custom Settings (Protected), or Salesforce Secrets'
        },
        {
            'rule': 'Generic Exception Catch',
            'category': 'Code Quality',
            'description': 'Catching generic Exception hides specific errors and makes debugging difficult',
            'severity': 'Low',
            'recommendation': 'Catch specific exception types (DmlException, QueryException, etc.)'
        },
        {
            'rule': 'System.debug with Sensitive Data',
            'category': 'Security',
            'description': 'Logging sensitive data (password, SSN, credit card) exposes it in debug logs',
            'severity': 'High',
            'recommendation': 'Remove sensitive data from System.debug() statements'
        },
        {
            'rule': 'Recursive Trigger Risk',
            'category': 'Code Quality',
            'description': 'Triggers without recursion guards can cause infinite loops',
            'severity': 'High',
            'recommendation': 'Implement static boolean flags or use trigger frameworks to prevent recursion'
        },
        {
            'rule': 'Nested Loops with DML/SOQL',
            'category': 'Performance',
            'description': 'Nested loops with SOQL/DML can exponentially increase governor limit consumption',
            'severity': 'Critical',
            'recommendation': 'Flatten nested loops or move SOQL/DML outside inner loops'
        },
        {
            'rule': 'Missing Test Assertions',
            'category': 'Test Quality',
            'description': 'Test methods without assertions do not validate expected behavior',
            'severity': 'Medium',
            'recommendation': 'Add System.assert(), System.assertEquals(), or Assert class methods'
        },
        {
            'rule': '@isTest(SeeAllData=true)',
            'category': 'Test Quality',
            'description': 'Tests using real org data are unreliable and can have different results across orgs',
            'severity': 'High',
            'recommendation': 'Remove SeeAllData=true and create test data using @testSetup or Test.startTest()'
        },
        {
            'rule': 'Missing Persona-Based Testing',
            'category': 'Test Quality',
            'description': 'Test classes without System.runAs() do not validate sharing rules, CRUD/FLS, or profile-specific behavior',
            'severity': 'Medium',
            'recommendation': 'Add System.runAs() to test with different user profiles/personas to ensure security and sharing work correctly'
        },
        {
            'rule': 'Test Coverage Below 90%',
            'category': 'Test Quality',
            'description': 'Classes with coverage from 80% up to but not including 90% should be improved to meet the recommended quality bar',
            'severity': 'Medium',
            'recommendation': 'Increase automated test coverage to at least 90%'
        },
        {
            'rule': 'Test Coverage Below 80%',
            'category': 'Test Quality',
            'description': 'Classes with coverage from 75% up to but not including 80% have elevated delivery risk and should be prioritized',
            'severity': 'High',
            'recommendation': 'Increase automated test coverage to at least 80% as a priority'
        },
        {
            'rule': 'Test Coverage Below 75%',
            'category': 'Test Quality',
            'description': 'Classes below 75% coverage are critically under-tested and fall below the minimum target',
            'severity': 'Critical',
            'recommendation': 'Increase automated test coverage to at least 75% immediately'
        },
        {
            'rule': 'Mixed DML Operations',
            'category': 'Architecture',
            'description': 'Setup objects (User, Profile) and standard objects cannot be modified in the same transaction',
            'severity': 'High',
            'recommendation': 'Use System.runAs() or separate the operations into different transactions'
        },
        {
            'rule': 'Schema.getGlobalDescribe() in Loop',
            'category': 'Performance',
            'description': 'Repeated Schema.getGlobalDescribe() calls inside loops are an Apex Guru anti-pattern that waste CPU and describe resources',
            'severity': 'High',
            'recommendation': 'Cache Schema.getGlobalDescribe() outside the loop and reuse the describe map'
        },
        {
            'rule': 'Schema.getGlobalDescribe() Not Efficient',
            'category': 'Performance',
            'description': 'Repeated uncached Schema.getGlobalDescribe() calls are inefficient when the describe map can be reused',
            'severity': 'Medium',
            'recommendation': 'Store Schema.getGlobalDescribe() in a local/static variable instead of calling it multiple times'
        },
        {
            'rule': 'Redundant SOQL',
            'category': 'Performance',
            'description': 'The same SOQL query is executed repeatedly instead of reusing previously fetched results',
            'severity': 'Medium',
            'recommendation': 'Reuse query results or cache them instead of issuing identical SOQL queries multiple times'
        },
        {
            'rule': 'SObject Map in a For Loop',
            'category': 'Performance',
            'description': 'Building an SObject map with map.put(record.Id, record) inside a loop is less efficient than using the built-in map constructor',
            'severity': 'Medium',
            'recommendation': 'Use new Map<Id, SObject>(records) instead of manually populating an SObject map in a loop'
        },
        {
            'rule': 'SOQL with Negative Expressions',
            'category': 'Performance',
            'description': 'Negative SOQL predicates such as !=, NOT IN, or NOT LIKE can reduce selectivity and hurt query performance',
            'severity': 'Medium',
            'recommendation': 'Prefer positive selective filters instead of negative SOQL expressions when possible'
        },
        {
            'rule': 'SOQL with Unused Fields',
            'category': 'Performance',
            'description': 'Selecting fields that are never referenced later increases heap usage and query cost unnecessarily',
            'severity': 'Medium',
            'recommendation': 'Remove unused fields from the SELECT clause when they are not needed by downstream logic'
        },
        {
            'rule': 'SOQL with Wildcard Filter',
            'category': 'Performance',
            'description': 'LIKE filters with wildcard patterns can become unselective and force expensive scans',
            'severity': 'Medium',
            'recommendation': 'Avoid broad wildcard LIKE filters when more selective exact or prefix filters are possible'
        },
        {
            'rule': 'SOQL Without a WHERE Clause or LIMIT Statement',
            'category': 'Performance',
            'description': 'SOQL queries without either a WHERE clause or LIMIT statement can retrieve excessive data',
            'severity': 'High',
            'recommendation': 'Add a WHERE clause or LIMIT statement to bound the result set'
        },
        {
            'rule': 'Unused Methods',
            'category': 'Best Practices',
            'description': 'Private methods that are never called add dead code and make classes harder to maintain',
            'severity': 'Low',
            'recommendation': 'Remove unused private methods or connect them to active code paths if still needed'
        },
        {
            'rule': 'SOQL with Apex Filter',
            'category': 'Performance',
            'description': 'Broad SOQL followed by filtering records in Apex is an Apex Guru anti-pattern that should be pushed into the WHERE clause',
            'severity': 'High',
            'recommendation': 'Move record filtering into SOQL WHERE clauses instead of looping and filtering in Apex'
        },
        {
            'rule': 'Expensive Methods in Loop',
            'category': 'Performance',
            'description': 'Expensive reflection/serialization/regex operations repeated inside loops consume unnecessary CPU',
            'severity': 'High',
            'recommendation': 'Evaluate expensive methods once outside the loop and reuse cached results'
        },
        {
            'rule': 'Expensive String Comparison',
            'category': 'Performance',
            'description': 'Repeated case-normalizing string comparisons create unnecessary allocations and CPU overhead',
            'severity': 'Medium',
            'recommendation': 'Normalize strings once or use more efficient comparison approaches like equalsIgnoreCase()'
        },
        {
            'rule': 'Copying Elements with for Loop',
            'category': 'Best Practices',
            'description': 'Element-by-element copying in a for loop is an Apex Guru anti-pattern when addAll() or direct collection reuse is possible',
            'severity': 'Medium',
            'recommendation': 'Replace manual collection copying loops with addAll() or direct collection assignment when no transformation is needed'
        },
        {
            'rule': 'Sorting in Apex Instead of SOQL ORDER BY',
            'category': 'Performance',
            'description': 'Sorting queried data in Apex instead of using SOQL ORDER BY is an Apex Guru anti-pattern for sortable database fields',
            'severity': 'Medium',
            'recommendation': 'Prefer ORDER BY in SOQL when the required ordering can be delegated to the database'
        },
        {
            'rule': 'Busy Loop Delay',
            'category': 'Performance',
            'description': 'Using loops as delays or spin-waits wastes CPU time and is an Apex Guru anti-pattern',
            'severity': 'High',
            'recommendation': 'Remove busy-wait loops and use asynchronous/event-driven approaches instead'
        },
        {
            'rule': 'Limits.getHeapSize() in Loop',
            'category': 'Performance',
            'description': 'Repeated Limits.getHeapSize()/getLimitHeapSize() checks inside loops add unnecessary overhead',
            'severity': 'Medium',
            'recommendation': 'Move heap-size monitoring outside loops or refactor to reduce heap pressure'
        }
    ]
    
    def __init__(self, config: Dict = None):
        """
        Initialize report generator
        
        Args:
            config: Configuration dictionary from config.yaml
        """
        # Start with defaults
        self.config = self._default_config()
        
        # Merge with provided config
        if config:
            self.config.update(config)
        
        # Colors
        self.colors = {
            'critical': 'FF0000',  # Red
            'high': 'FFA500',      # Orange
            'medium': 'FFFF00',    # Yellow
            'low': 'ADD8E6',       # Light Blue
            'false_positive': 'D9D9D9',  # Light Gray
            'excellent': '90EE90', # Light Green
            'dark_green': '006400', # Dark Green
            'header': '4472C4',    # Blue
            'good': 'C6E0B4',      # Light Green
            'actionable_header': 'C55A11',  # Brown/Orange
            'remarks_fill': 'FFF2CC',  # Light Yellow
            'status_fill': 'E2F0D9',  # Light Green
        }
    
    def _default_config(self) -> Dict:
        """Default report configuration"""
        return {
            'directory': './audit_reports',
            'filename_template': 'SF_Audit_{org_name}_{timestamp}.xlsx',
            'sheets': {
                'metadata': 'Metadata Snapshot',
                'executive': 'Executive Summary',
                'detailed': 'Detailed Audit',
                'coverage': 'Code Coverage Detail'
            },
            'highlight_critical': True,
            'color_coding': True,
            'freeze_panes': True,
            'auto_filter': True
        }

    @classmethod
    def _get_rule_source(cls, rule_name: str) -> str:
        """Return the source label for a rule"""
        return 'Apex Guru' if rule_name in cls.APEX_GURU_RULES else 'Audit Tool'
    
    def _add_grade_conditional_formatting(self, ws, cell_ref):
        """Add conditional formatting rules so grade cell color updates dynamically"""
        grade_formats = [
            ('"Excellent"', '90EE90', '000000'),
            ('"Very Good"', 'C6E0B4', '000000'),
            ('"Good"', 'D3D3D3', '000000'),
            ('"Average"', 'FFFF00', '000000'),
            ('"Below Average"', 'FF0000', 'FFFFFF'),
            ('"Poor"', '8B0000', 'FFFFFF'),
        ]
        for formula_val, bg_color, font_color in grade_formats:
            ws.conditional_formatting.add(cell_ref,
                CellIsRule(
                    operator='equal',
                    formula=[formula_val],
                    fill=PatternFill(start_color=bg_color, fill_type='solid'),
                    font=Font(size=14, bold=True, color=font_color)
                ))
    
    def generate_report(
        self,
        org_name: str,
        grading_result: GradingResult,
        metadata_snapshot: Dict,
        detailed_issues: List[Dict],
        data_model_issues: List[Dict],
        coverage_details: List[Dict],
        all_violations: List[Any] = None
    ) -> str:
        """
        Generate complete audit report
        
        Args:
            org_name: Name of the Salesforce org
            grading_result: Grading result from GradingEngine
            metadata_snapshot: Org metadata stats
            detailed_issues: All issues found
            data_model_issues: Data model specific issues
            coverage_details: Test coverage details
            all_violations: All code violations
        
        Returns:
            Path to generated report file
        """
        # Create output directory
        os.makedirs(self.config['directory'], exist_ok=True)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = self.config['filename_template'].format(
            org_name=org_name.replace(' ', '_'),
            timestamp=timestamp
        )
        filepath = os.path.join(self.config['directory'], filename)
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Generate each sheet
        self._create_metadata_sheet(wb, metadata_snapshot, grading_result)
        self._create_executive_sheet(wb, grading_result, detailed_issues)
        
        # Keep Grading Calculator as the 3rd sheet
        self._create_grading_calculator_sheet(wb, grading_result)
        
        self._create_detailed_sheet(wb, detailed_issues)
        
        # Conditionally create data model sheet
        if data_model_issues:
            self._create_data_model_sheet(wb, data_model_issues)
        
        # Conditionally create LWC sheet
        lwc_issues = [issue for issue in detailed_issues if issue.get('type') == 'LWC']
        if lwc_issues:
            self._create_lwc_sheet(wb, lwc_issues)
        
        self._create_coverage_sheet(wb, coverage_details)
        
        # Add Rules Reference Sheet (always last)
        self._create_rules_reference_sheet(wb)
        
        # Save workbook
        wb.save(filepath)
        
        print(f"✅ Report generated: {filepath}")
        return filepath
    
    def _create_metadata_sheet(
        self,
        wb: Workbook,
        metadata: Dict,
        grading: GradingResult
    ):
        """Create Sheet 0 - Metadata Snapshot"""
        ws = wb.create_sheet(self.config['sheets']['metadata'])
        
        # Title
        ws['A1'] = 'Salesforce Code Audit - Metadata Snapshot'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color=self.colors['header'], 
                                     fill_type='solid')
        ws.merge_cells('A1:B1')

        def add_section_header(title: str, target_row: int, fill_color: str = None):
            fill = fill_color or self.colors['header']
            ws[f'A{target_row}'] = title
            ws[f'A{target_row}'].font = Font(size=12, bold=True, color='FFFFFF')
            ws[f'A{target_row}'].fill = PatternFill(start_color=fill, fill_type='solid')
            ws[f'A{target_row}'].alignment = Alignment(horizontal='center')
            ws.merge_cells(f'A{target_row}:B{target_row}')
        
        # Audit info
        row = 3
        audit_info = [
            ('Audit Timestamp', metadata.get('audit_timestamp', datetime.now().isoformat())),
            ('Org Name', metadata.get('org_name', 'N/A')),
            ('Org Type', metadata.get('org_type', 'N/A')),
            ('API Version', metadata.get('api_version', 'N/A')),
            ('Org Domain Name', metadata.get('org_domain_name', 'N/A')),
        ]
        
        grade_row = None
        for label, value in audit_info:
            if label:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                
                # Track grade row for conditional formatting
                if label == 'Overall Code Health Grade':
                    grade_row = row
                    ws[f'B{row}'].font = Font(size=14, bold=True)
            row += 1
        
        row += 1
        add_section_header('Overall Code Health Grade', row)
        grade_row = row + 1
        ws[f'A{grade_row}'] = 'Current Grade'
        ws[f'A{grade_row}'].font = Font(bold=True)
        ws[f'B{grade_row}'] = "='Grading Calculator'!B18"
        ws[f'B{grade_row}'].font = Font(size=14, bold=True)
        self._add_grade_conditional_formatting(ws, f'B{grade_row}')
        
        # Counts section
        row = grade_row + 2
        add_section_header('Org Statistics', row)
        row += 1
        
        counts = [
            ('Apex Classes', metadata.get('apex_classes', 0)),
            ('Test Classes', metadata.get('test_classes', 0)),
            ('Triggers', metadata.get('triggers', 0)),
            ('Custom Objects', metadata.get('custom_objects', 0)),
        ]
        
        # Add flows if found
        if metadata.get('flows', 0) > 0:
            counts.insert(4, ('Flows', metadata.get('flows', 0)))
        
        for label, value in counts:
            if label:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
            row += 1

        add_section_header('Test Coverage (%)', row)
        coverage_row = row + 1
        ws[f'A{coverage_row}'] = 'Current Coverage'
        ws[f'A{coverage_row}'].font = Font(bold=True)
        ws[f'B{coverage_row}'] = "='Grading Calculator'!B17"
        ws[f'B{coverage_row}'].font = Font(bold=True, color='FFFFFF')
        ws[f'B{coverage_row}'].number_format = '0.0"%"'
        ws.conditional_formatting.add(
            f'B{coverage_row}',
            CellIsRule(
                operator='lessThan',
                formula=['75'],
                fill=PatternFill(start_color=self.colors['critical'], fill_type='solid'),
                font=Font(bold=True, color='FFFFFF')
            )
        )
        ws.conditional_formatting.add(
            f'B{coverage_row}',
            CellIsRule(
                operator='greaterThan',
                formula=['75'],
                fill=PatternFill(start_color=self.colors['dark_green'], fill_type='solid'),
                font=Font(bold=True, color='FFFFFF')
            )
        )
        
        # Issues summary
        row = coverage_row + 2
        add_section_header('Issue Summary', row)
        row += 1
        
        issues = [
            ('Critical Issues', "='Grading Calculator'!B7", self.colors['critical']),
            ('High Issues', "='Grading Calculator'!B8", self.colors['high']),
            ('Medium Issues', "='Grading Calculator'!B9", self.colors['medium']),
            ('Low Issues', "='Grading Calculator'!B10", self.colors['low']),
            ('Total Issues', "='Grading Calculator'!B11", None),
        ]
        
        for label, value, color in issues:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            if color:
                ws[f'A{row}'].fill = PatternFill(start_color=color, fill_type='solid')
            if color:
                ws[f'B{row}'].fill = PatternFill(start_color=color, fill_type='solid')
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
    
    def _create_executive_sheet(
        self,
        wb: Workbook,
        grading: GradingResult,
        all_issues: List[Dict]
    ):
        """Create Sheet 1 - Executive Summary"""
        ws = wb.create_sheet(self.config['sheets']['executive'])
        
        # Title
        ws['A1'] = 'Executive Summary'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        ws.merge_cells('A1:E1')
        
        # Grade (prominent) - formula linked to Grading Calculator for dynamic updates
        ws['A3'] = 'Code Health Grade:'
        ws['A3'].font = Font(size=14, bold=True)
        ws['B3'] = "='Grading Calculator'!B18"
        ws['B3'].font = Font(size=14, bold=True)
        
        # Dynamic color coding via conditional formatting (updates when grade changes)
        self._add_grade_conditional_formatting(ws, 'B3')
        
        # Executive Opinion
        row = 5
        ws[f'A{row}'] = 'Executive Opinion'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        ws[f'A{row}'] = grading.executive_summary
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[row].height = 100
        
        # Top Priority Fixes
        row += 2
        ws[f'A{row}'] = 'Top 5 Priority Fixes'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        # Headers
        headers = ['Rule/Issue', 'File', 'Line', 'Criticality', 'Recommendation']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        row += 1
        
        # Top fixes data
        for fix in grading.top_priority_fixes[:5]:
            ws.cell(row=row, column=1).value = fix.get('rule', 'N/A')
            ws.cell(row=row, column=2).value = fix.get('file', 'N/A')
            ws.cell(row=row, column=3).value = fix.get('line', 'N/A')
            ws.cell(row=row, column=4).value = fix.get('criticality', 'N/A')
            ws.cell(row=row, column=5).value = fix.get('recommendation', 'N/A')
            
            # Color code criticality
            criticality = fix.get('criticality', '').lower()
            if criticality in self.colors:
                ws.cell(row=row, column=4).fill = PatternFill(
                    start_color=self.colors[criticality],
                    fill_type='solid'
                )
            row += 1
        
        # Issue Pivot
        row += 2
        ws[f'A{row}'] = 'Issues by Category and Criticality'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        # Pivot headers
        pivot_headers = ['Category', 'Critical', 'High', 'Medium', 'Low', 'Total']
        for col, header in enumerate(pivot_headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        row += 1
        
        # Generate pivot data
        from grading_engine import GradingEngine
        engine = GradingEngine()
        pivot = engine.generate_issue_pivot(all_issues)
        
        for category, counts in pivot.items():
            ws.cell(row=row, column=1).value = category
            ws.cell(row=row, column=2).value = counts.get('Critical', 0)
            ws.cell(row=row, column=3).value = counts.get('High', 0)
            ws.cell(row=row, column=4).value = counts.get('Medium', 0)
            ws.cell(row=row, column=5).value = counts.get('Low', 0)
            ws.cell(row=row, column=6).value = sum(counts.values())
            
            # Color code cells
            if counts.get('Critical', 0) > 0:
                ws.cell(row=row, column=2).fill = PatternFill(
                    start_color=self.colors['critical'], fill_type='solid'
                )
            if counts.get('High', 0) > 0:
                ws.cell(row=row, column=3).fill = PatternFill(
                    start_color=self.colors['high'], fill_type='solid'
                )
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 50
        
        # Freeze panes
        if self.config.get('freeze_panes', True):
            ws.freeze_panes = 'A2'
    
    def _create_detailed_sheet(self, wb: Workbook, issues: List[Dict]):
        """Create Sheet 2 - Detailed Audit"""
        ws = wb.create_sheet(self.config['sheets']['detailed'])
        
        # Title
        ws['A1'] = 'Detailed Audit Results'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:K1')
        
        # Headers
        headers = [
            'Rule Source',
            'File/Object Name',
            'Type',
            'Line Number',
            'Category',
            'Rule/Issue Name',
            'Criticality',
            'Snippet/Evidence',
            'Recommendation',
            'Architect/Dev Remarks',
            'Fix Status'
        ]
        
        row = 2
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            if header in ['Architect/Dev Remarks', 'Fix Status']:
                cell.fill = PatternFill(start_color=self.colors['actionable_header'], fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
            else:
                cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        
        # Data rows
        row = 3
        for issue in issues:
            ws.cell(row=row, column=1).value = issue.get('rule_source', self._get_rule_source(issue.get('rule_name', '')))
            ws.cell(row=row, column=2).value = issue.get('file_name', 'N/A')
            ws.cell(row=row, column=3).value = issue.get('type', 'N/A')
            ws.cell(row=row, column=4).value = issue.get('line_number', 'N/A')
            ws.cell(row=row, column=5).value = issue.get('category', 'N/A')
            ws.cell(row=row, column=6).value = issue.get('rule_name', 'N/A')
            ws.cell(row=row, column=7).value = issue.get('criticality', 'N/A')
            ws.cell(row=row, column=8).value = issue.get('snippet', 'N/A')
            ws.cell(row=row, column=9).value = issue.get('recommendation', 'N/A')
            ws.cell(row=row, column=10).value = issue.get('architect_dev_remarks', '')
            ws.cell(row=row, column=11).value = issue.get('fix_status', 'Not Started')
            
            # Color code by criticality
            criticality = issue.get('criticality', '').lower().replace(' ', '_')
            if criticality in self.colors and self.config.get('color_coding', True):
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color=self.colors[criticality],
                        fill_type='solid'
                    )
            
            # Keep actionable columns visually distinct from severity coloring
            ws.cell(row=row, column=10).fill = PatternFill(
                start_color=self.colors['remarks_fill'],
                fill_type='solid'
            )
            ws.cell(row=row, column=11).fill = PatternFill(
                start_color=self.colors['status_fill'],
                fill_type='solid'
            )
            
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 50
        ws.column_dimensions['I'].width = 50
        ws.column_dimensions['J'].width = 35
        ws.column_dimensions['K'].width = 18
        
        # Auto filter
        if self.config.get('auto_filter', True):
            ws.auto_filter.ref = f'A2:K{row-1}'
        
        # Data validation dropdown for Criticality column (G)
        # Allows users to change severity - grade auto-recalculates
        dv = DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low,False Positive"',
            allow_blank=True
        )
        dv.error = "Please select: Critical, High, Medium, Low, or False Positive"
        dv.errorTitle = "Invalid Criticality"
        dv.prompt = "Select criticality level - grade auto-updates"
        dv.promptTitle = "Criticality"
        ws.add_data_validation(dv)
        dv.add(f'G3:G{max(row - 1, 3)}')

        # Data validation dropdown for Fix Status column (K)
        status_dv = DataValidation(
            type="list",
            formula1='"Fixed,In Progress,Not Started,False Positive"',
            allow_blank=True
        )
        status_dv.error = "Please select: Fixed, In Progress, Not Started, or False Positive"
        status_dv.errorTitle = "Invalid Fix Status"
        status_dv.prompt = "Select fix status"
        status_dv.promptTitle = "Fix Status"
        ws.add_data_validation(status_dv)
        status_dv.add(f'K3:K{max(row - 1, 3)}')
        
        # Conditional formatting for dynamic row colors based on Criticality
        # When user changes criticality via dropdown, row color updates automatically
        fmt_range = f'A3:I{max(row - 1, 3)}'
        ws.conditional_formatting.add(fmt_range,
            FormulaRule(formula=['$G3="Critical"'],
                       fill=PatternFill(start_color=self.colors['critical'], fill_type='solid')))
        ws.conditional_formatting.add(fmt_range,
            FormulaRule(formula=['$G3="High"'],
                       fill=PatternFill(start_color=self.colors['high'], fill_type='solid')))
        ws.conditional_formatting.add(fmt_range,
            FormulaRule(formula=['$G3="Medium"'],
                       fill=PatternFill(start_color=self.colors['medium'], fill_type='solid')))
        ws.conditional_formatting.add(fmt_range,
            FormulaRule(formula=['$G3="Low"'],
                       fill=PatternFill(start_color=self.colors['low'], fill_type='solid')))
        ws.conditional_formatting.add(fmt_range,
            FormulaRule(formula=['$G3="False Positive"'],
                       fill=PatternFill(start_color=self.colors['false_positive'], fill_type='solid')))
        
        # Freeze panes
        if self.config.get('freeze_panes', True):
            ws.freeze_panes = 'A3'
    
    def _create_data_model_sheet(self, wb: Workbook, issues: List[Dict]):
        """Create Sheet 3 - Data Model Issues"""
        ws = wb.create_sheet(self.config['sheets']['data_model'])
        
        # Title
        ws['A1'] = 'Data Model Issues'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        # Headers
        headers = ['Object Name', 'Field Name', 'Issue Type', 'Impact', 'Recommendation']
        row = 2
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        
        # Data rows
        row = 3
        for issue in issues:
            ws.cell(row=row, column=1).value = issue.get('object_name', 'N/A')
            ws.cell(row=row, column=2).value = issue.get('field_name', 'N/A')
            ws.cell(row=row, column=3).value = issue.get('issue_type', 'N/A')
            ws.cell(row=row, column=4).value = issue.get('impact', 'N/A')
            ws.cell(row=row, column=5).value = issue.get('recommendation', 'N/A')
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 40
        
        # Auto filter
        if self.config.get('auto_filter', True) and row > 3:
            ws.auto_filter.ref = f'A2:E{row-1}'
        
        # Freeze panes
        if self.config.get('freeze_panes', True):
            ws.freeze_panes = 'A3'
    
    def _create_lwc_sheet(self, wb: Workbook, issues: List[Dict]):
        """Create LWC Security & Best Practices Sheet"""
        ws = wb.create_sheet('LWC Security & Best Practices')
        
        # Title
        ws['A1'] = 'LWC Security & Best Practices Issues'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:G1')
        
        # Headers
        headers = [
            'Component Name',
            'File',
            'Line',
            'Category',
            'Issue',
            'Criticality',
            'Recommendation'
        ]
        
        row = 2
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        
        # Data rows
        row = 3
        for issue in issues:
            ws.cell(row=row, column=1).value = issue.get('file_name', 'N/A').replace('.js', '').replace('.html', '')
            ws.cell(row=row, column=2).value = issue.get('file_name', 'N/A')
            ws.cell(row=row, column=3).value = issue.get('line_number', 'N/A')
            ws.cell(row=row, column=4).value = issue.get('category', 'N/A')
            ws.cell(row=row, column=5).value = issue.get('rule_name', 'N/A')
            
            # Criticality with color
            criticality = issue.get('criticality', 'Low')
            cell = ws.cell(row=row, column=6)
            cell.value = criticality
            cell.fill = PatternFill(
                start_color=self.colors.get(criticality.lower(), 'FFFFFF'),
                fill_type='solid'
            )
            
            ws.cell(row=row, column=7).value = issue.get('recommendation', 'N/A')
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 50
        
        # Auto filter
        if self.config.get('auto_filter', True) and row > 3:
            ws.auto_filter.ref = f'A2:G{row-1}'
        
        # Freeze panes
        if self.config.get('freeze_panes', True):
            ws.freeze_panes = 'A3'
    
    def _create_coverage_sheet(self, wb: Workbook, coverage_data: List[Dict]):
        """Create Sheet 4 - Code Coverage Detail"""
        ws = wb.create_sheet(self.config['sheets']['coverage'])
        
        # Title
        ws['A1'] = 'Code Coverage Detail'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Headers
        headers = [
            'Class/Trigger Name',
            'Type',
            'Coverage %',
            'Lines Covered',
            'Lines Not Covered',
            'Test Methods'
        ]
        
        row = 2
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        
        # Data rows
        row = 3
        
        if not coverage_data or len(coverage_data) == 0:
            # Add a message when there's no coverage data
            ws.merge_cells('A3:F3')
            cell = ws.cell(row=3, column=1)
            cell.value = 'ℹ️  No test coverage data found. Tests may not have been run in this org yet.'
            cell.font = Font(size=12, italic=True, color='FF6600')
            cell.fill = PatternFill(start_color='FFF4E6', fill_type='solid')
            row = 4
        else:
            for item in coverage_data:
                ws.cell(row=row, column=1).value = item.get('name', 'N/A')
                ws.cell(row=row, column=2).value = item.get('type', 'N/A')
                ws.cell(row=row, column=3).value = f"{item.get('coverage_percent', 0):.1f}%"
                ws.cell(row=row, column=4).value = item.get('lines_covered', 0)
                ws.cell(row=row, column=5).value = item.get('lines_not_covered', 0)
                ws.cell(row=row, column=6).value = item.get('test_methods', 'N/A')
                
                # Highlight rows with coverage < 90%
                coverage = item.get('coverage_percent', 100)
                if coverage < 90 and self.config.get('highlight_critical', True):
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color='FFE6E6',  # Light red
                            fill_type='solid'
                        )
                        ws.cell(row=row, column=col).font = Font(color='FF0000', bold=True)
                
                row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 20
        
        # Auto filter
        if self.config.get('auto_filter', True) and row > 3:
            ws.auto_filter.ref = f'A2:F{row-1}'
        
        # Freeze panes
        if self.config.get('freeze_panes', True):
            ws.freeze_panes = 'A3'
    
    def _create_grading_calculator_sheet(self, wb, grading):
        """Create Grading Calculator sheet with dynamic Excel formulas.
        
        When the user changes issue severity in the Detailed Audit sheet,
        the grade automatically recalculates here and propagates to
        Metadata Snapshot and Executive Summary sheets.
        
        Cell Reference Map:
            B7  = Critical issue count (COUNTIF)
            B8  = High issue count (COUNTIF)
            B9  = Medium issue count (COUNTIF)
            B10 = Low issue count (COUNTIF)
            B11 = Total issue count (SUM)
            B13 = SOQL/DML in loops detected (YES/NO)
            B14 = Loop violation count
            B16 = Org test coverage % (editable)
            B18 = CALCULATED GRADE (nested IF waterfall)
            B19 = Grade rationale (nested IF)
        """
        ws = wb.create_sheet('Grading Calculator')
        detailed_sheet = self.config['sheets']['detailed']
        
        # ── Title ──
        ws['A1'] = 'Dynamic Grading Calculator'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        ws.merge_cells('A1:D1')
        
        ws['A2'] = 'Change issue severity in "Detailed Audit" column G → Grade auto-updates below'
        ws['A2'].font = Font(size=11, italic=True, color='0000FF')
        ws.merge_cells('A2:D2')
        ws['A3'] = 'You can also adjust "Org Coverage %" in B16 to simulate grade impact'
        ws['A3'].font = Font(size=11, italic=True, color='0000FF')
        ws.merge_cells('A3:D3')
        
        # ── Issue Counts Section (rows 5-11) ──
        ws['A5'] = 'Issue Counts (Auto-calculated from Detailed Audit)'
        ws['A5'].font = Font(size=13, bold=True)
        ws.merge_cells('A5:C5')
        
        # Column headers row 6
        for col, header in [('A', 'Severity'), ('B', 'Count'), ('C', 'Source')]:
            ws[f'{col}6'] = header
            ws[f'{col}6'].font = Font(bold=True, color='FFFFFF')
            ws[f'{col}6'].fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
        
        # B7: Critical count
        ws['A7'] = 'Critical'
        ws['A7'].font = Font(bold=True)
        ws['B7'] = f"=COUNTIF('{detailed_sheet}'!G:G,\"Critical\")"
        ws['B7'].fill = PatternFill(start_color=self.colors['critical'], fill_type='solid')
        ws['B7'].font = Font(bold=True, size=13)
        ws['C7'] = 'Auto-counted from Detailed Audit column G'
        ws['C7'].font = Font(italic=True, color='888888', size=9)
        
        # B8: High count
        ws['A8'] = 'High'
        ws['A8'].font = Font(bold=True)
        ws['B8'] = f"=COUNTIF('{detailed_sheet}'!G:G,\"High\")"
        ws['B8'].fill = PatternFill(start_color=self.colors['high'], fill_type='solid')
        ws['B8'].font = Font(bold=True, size=13)
        ws['C8'] = 'Auto-counted from Detailed Audit column G'
        ws['C8'].font = Font(italic=True, color='888888', size=9)
        
        # B9: Medium count
        ws['A9'] = 'Medium'
        ws['A9'].font = Font(bold=True)
        ws['B9'] = f"=COUNTIF('{detailed_sheet}'!G:G,\"Medium\")"
        ws['B9'].fill = PatternFill(start_color=self.colors['medium'], fill_type='solid')
        ws['B9'].font = Font(bold=True, size=13)
        ws['C9'] = 'Auto-counted from Detailed Audit column G'
        ws['C9'].font = Font(italic=True, color='888888', size=9)
        
        # B10: Low count
        ws['A10'] = 'Low'
        ws['A10'].font = Font(bold=True)
        ws['B10'] = f"=COUNTIF('{detailed_sheet}'!G:G,\"Low\")"
        ws['B10'].fill = PatternFill(start_color=self.colors['low'], fill_type='solid')
        ws['B10'].font = Font(bold=True, size=13)
        ws['C10'] = 'Auto-counted from Detailed Audit column G'
        ws['C10'].font = Font(italic=True, color='888888', size=9)
        
        # B11: Total
        ws['A11'] = 'Total'
        ws['A11'].font = Font(bold=True, size=12)
        ws['B11'] = '=SUM(B7:B10)'
        ws['B11'].font = Font(bold=True, size=13)
        ws['B11'].border = Border(top=Side(style='double'))
        
        # ── SOQL/DML in Loops Detection (rows 13-14) ──
        ws['A13'] = 'SOQL/DML in Loops Detected'
        ws['A13'].font = Font(bold=True, size=12)
        loop_count_formula = (
            f'=SUM('
            f'COUNTIFS(\'{detailed_sheet}\'!F:F,"*SOQL*Loop*",\'{detailed_sheet}\'!G:G,"<>False Positive",\'{detailed_sheet}\'!K:K,"<>False Positive"),'
            f'COUNTIFS(\'{detailed_sheet}\'!F:F,"*DML*Loop*",\'{detailed_sheet}\'!G:G,"<>False Positive",\'{detailed_sheet}\'!K:K,"<>False Positive"),'
            f'COUNTIFS(\'{detailed_sheet}\'!F:F,"*SOQL_IN_LOOP*",\'{detailed_sheet}\'!G:G,"<>False Positive",\'{detailed_sheet}\'!K:K,"<>False Positive"),'
            f'COUNTIFS(\'{detailed_sheet}\'!F:F,"*DML_IN_LOOP*",\'{detailed_sheet}\'!G:G,"<>False Positive",\'{detailed_sheet}\'!K:K,"<>False Positive")'
            f')'
        )
        loops_formula = (
            f'=IF(B14>0,"YES","NO")'
        )
        ws['B13'] = loops_formula
        ws['B13'].font = Font(bold=True, size=13)
        ws['C13'] = 'Auto-detected from Detailed Audit; rows marked False Positive in Criticality or Fix Status are ignored'
        ws['C13'].font = Font(italic=True, color='888888', size=9)
        
        # Conditional format: YES = red, NO = green
        ws.conditional_formatting.add('B13',
            CellIsRule(operator='equal', formula=['"YES"'],
                      fill=PatternFill(start_color='FF0000', fill_type='solid'),
                      font=Font(bold=True, size=13, color='FFFFFF')))
        ws.conditional_formatting.add('B13',
            CellIsRule(operator='equal', formula=['"NO"'],
                      fill=PatternFill(start_color='90EE90', fill_type='solid'),
                      font=Font(bold=True, size=13)))
        
        ws['A14'] = 'Loop Violation Count'
        ws['A14'].font = Font(bold=True)
        ws['B14'] = loop_count_formula
        
        # ── Org Coverage - Editable (row 16) ──
        ws['A16'] = 'Org Test Coverage (%)'
        ws['A16'].font = Font(bold=True, size=13)
        ws['B16'] = grading.coverage_stats.org_coverage
        ws['B16'].font = Font(bold=True, size=16)
        ws['B16'].number_format = '0.0'
        ws['B16'].fill = PatternFill(start_color='E6F3FF', fill_type='solid')
        ws['B16'].border = Border(
            left=Side(style='medium', color='0000FF'),
            right=Side(style='medium', color='0000FF'),
            top=Side(style='medium', color='0000FF'),
            bottom=Side(style='medium', color='0000FF'))
        ws['C16'] = 'EDITABLE: Enter 80 or 80% to simulate grade impact'
        ws['C16'].font = Font(size=11, italic=True, color='008000')

        # Normalize user input so both 80 and 80% work correctly
        ws['A17'] = 'Normalized Coverage (%)'
        ws['A17'].font = Font(bold=True)
        ws['B17'] = '=IF(B16<=1,B16*100,B16)'
        ws['B17'].font = Font(size=12, bold=True, color='666666')
        ws['B17'].number_format = '0.0'
        ws['C17'] = 'Auto-normalized for grading formulas'
        ws['C17'].font = Font(size=10, italic=True, color='888888')
        
        # ── CALCULATED GRADE (row 18) ──
        ws['A18'] = 'CALCULATED GRADE'
        ws['A18'].font = Font(size=16, bold=True)
        
        # Nested IF formula implementing the waterfall grading logic:
        # Poor → Excellent → Very Good → Good → Average → Below Average
        grade_formula = (
            '=IF(OR(B13="YES",B7>=10,B17<50),"Poor",'
            'IF(AND(B17>=90,B7=0,B8=0,B9<15),"Excellent",'
            'IF(AND(B17>=85,B7=0,B8<10,B9<30),"Very Good",'
            'IF(AND(B17>=75,B7<=3,B8<15),"Good",'
            'IF(AND(B17>=70,B7<=5,B8<50),"Average",'
            'IF(AND(B17>=50,B7<=10),"Below Average","Poor"))))))'
        )
        ws['B18'] = grade_formula
        ws['B18'].font = Font(size=18, bold=True)
        self._add_grade_conditional_formatting(ws, 'B18')
        
        # ── Grade Rationale (row 19) ──
        ws['A19'] = 'Grade Rationale'
        ws['A19'].font = Font(bold=True, size=11)
        rationale_formula = (
            '=IF(B13="YES","Severe: SOQL/DML in loops detected, which falls into the Poor grade band",'
            'IF(B7>=10,"Severe: "&B7&" Critical issues found (10+), major refactoring required",'
            'IF(B17<50,"Severe: Coverage "&TEXT(B17,"0.0")&"% (below 50%)",'
            'IF(AND(B17>=90,B7=0,B8=0,B9<15),"No Critical or High issues, "&B9&" Medium issues, "&TEXT(B17,"0.0")&"% coverage",'
            'IF(AND(B17>=85,B7=0,B8<10,B9<30),"0 Critical, "&B8&" High, "&B9&" Medium, "&TEXT(B17,"0.0")&"% coverage",'
            'IF(AND(B17>=75,B7<=3,B8<15),B7&" Critical, "&B8&" High, "&TEXT(B17,"0.0")&"% coverage meets Good band",'
            'IF(AND(B17>=70,B7<=5,B8<50),B7&" Critical, "&B8&" High, "&TEXT(B17,"0.0")&"% coverage places this in Average",'
            'IF(AND(B17>=50,B7<=10),B7&" Critical issues and "&TEXT(B17,"0.0")&"% coverage place this in Below Average",'
            '"Does not meet minimum grading thresholds"))))))))'
        )
        ws['B19'] = rationale_formula
        ws['B19'].alignment = Alignment(wrap_text=True)
        ws.merge_cells('B19:D19')
        
        # ── Why This Grade? (rows 21-28) ──
        ws['A21'] = 'Why This Grade?'
        ws['A21'].font = Font(size=13, bold=True)
        ws.merge_cells('A21:D21')

        for col_idx, header in enumerate(['Factor', 'Current Value', 'Status', 'What It Means'], 1):
            cell = ws.cell(row=22, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')

        breakdown_rows = [
            (
                23,
                'Loop Rule',
                '=B13',
                '=IF(B13="YES","BLOCKER","OK")',
                '=IF(B13="YES","SOQL/DML in loops forces the grade to Poor","No loop blocker")'
            ),
            (
                24,
                'Coverage Band',
                '=TEXT(B17,"0.0")&"%"',
                '=IF(B17<50,"POOR BAND",IF(B17<70,"BELOW AVG BAND",IF(B17<75,"AVERAGE BAND",IF(B17<85,"GOOD BAND",IF(B17<90,"VERY GOOD BAND","EXCELLENT BAND")))))',
                '=IF(B17<50,"Coverage alone limits the grade to Poor",IF(B17<70,"Coverage cannot reach Average or better",IF(B17<75,"Coverage cannot reach Good or better",IF(B17<85,"Coverage cannot reach Very Good or Excellent",IF(B17<90,"Coverage cannot reach Excellent","Coverage supports Excellent")))))'
            ),
            (
                25,
                'Critical Issues',
                '=B7',
                '=IF(B7>=10,"POOR BLOCKER",IF(B7>5,"BELOW AVG CAP",IF(B7>3,"AVERAGE CAP",IF(B7>0,"GOOD CAP","TOP BAND READY"))))',
                '=IF(B7>=10,"10 or more Critical issues forces Poor",IF(B7>5,"More than 5 Critical issues prevents Average or better",IF(B7>3,"More than 3 Critical issues prevents Good or better",IF(B7>0,"Any Critical issues prevents Very Good or Excellent","Critical count supports top bands"))))'
            ),
            (
                26,
                'High Issues',
                '=B8',
                '=IF(B8>=50,"AVERAGE BLOCKER",IF(B8>=15,"GOOD BLOCKER",IF(B8>=10,"VERY GOOD BLOCKER",IF(B8>0,"EXCELLENT BLOCKER","TOP BAND READY"))))',
                '=IF(B8>=50,"50 or more High issues prevents Average or better",IF(B8>=15,"15 or more High issues prevents Good or better",IF(B8>=10,"10 or more High issues prevents Very Good",IF(B8>0,"Any High issues prevents Excellent","High count supports Excellent"))))'
            ),
            (
                27,
                'Medium Issues',
                '=B9',
                '=IF(B9>=30,"VERY GOOD BLOCKER",IF(B9>=15,"EXCELLENT BLOCKER","TOP BAND READY"))',
                '=IF(B9>=30,"30 or more Medium issues prevents Very Good or Excellent",IF(B9>=15,"15 or more Medium issues prevents Excellent","Medium count supports Excellent"))'
            ),
            (
                28,
                'Next Better Grade Blocked By',
                '=B18',
                'NEXT STEP',
                '=IF(B18="Poor","Fix the Poor-grade blocker first (loops, coverage < 50%, or 10+ Critical issues)",IF(B18="Below Average","Raise coverage to at least 70% and reduce Critical issues to 5 or fewer",IF(B18="Average","Raise coverage to at least 75%, reduce Critical issues to 3 or fewer, and keep High issues below 50",IF(B18="Good","Raise coverage to at least 85%, reduce Critical issues to 0, keep High issues below 10, and keep Medium issues below 30",IF(B18="Very Good","Raise coverage to at least 90%, reduce High issues to 0, and keep Medium issues below 15","Already at highest grade")))))'
            ),
        ]

        for r, label, current_value, status_formula, meaning_formula in breakdown_rows:
            ws.cell(row=r, column=1).value = label
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.cell(row=r, column=2).value = current_value
            ws.cell(row=r, column=3).value = status_formula
            ws.cell(row=r, column=4).value = meaning_formula
            ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True)
            for col in range(1, 5):
                ws.cell(row=r, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

        # ── Grade Scale Reference Table (rows 30-37) ──
        ws['A30'] = 'Grade Scale Reference'
        ws['A30'].font = Font(size=13, bold=True)
        ws.merge_cells('A30:D30')

        # Headers row 31
        for col_idx, header in enumerate(['Grade', 'Coverage Threshold', 'Critical Issues', 'Other Criteria'], 1):
            cell = ws.cell(row=31, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')

        grade_ref = [
            ('Excellent', '>= 90%', '0', '0 High, < 15 Medium', '90EE90', '000000'),
            ('Very Good', '>= 85%', '0', '< 10 High issues, < 30 Medium', 'C6E0B4', '000000'),
            ('Good', '>= 75%', '<= 3', 'May have < 15 High', 'D3D3D3', '000000'),
            ('Average', '70-75%', '<= 5', '< 50 High', 'FFFF00', '000000'),
            ('Below Average', '50-70%', '<= 10', 'OR any SOQL/DML in loops', 'FF0000', 'FFFFFF'),
            ('Poor', '< 50%', '>= 10', 'OR loops', '8B0000', 'FFFFFF'),
        ]

        for idx, (grade_name, coverage, critical, other, bg_color, font_color) in enumerate(grade_ref):
            r = 32 + idx
            ws.cell(row=r, column=1).value = grade_name
            ws.cell(row=r, column=1).font = Font(bold=True, color=font_color)
            ws.cell(row=r, column=1).fill = PatternFill(start_color=bg_color, fill_type='solid')
            ws.cell(row=r, column=2).value = coverage
            ws.cell(row=r, column=3).value = critical
            ws.cell(row=r, column=4).value = other
            for col in range(1, 5):
                ws.cell(row=r, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

        # ── How to Use (rows 40+) ──
        ws['A40'] = 'How to Use This Sheet'
        ws['A40'].font = Font(size=13, bold=True)
        ws.merge_cells('A40:D40')
        
        instructions = [
            '1. Go to "Detailed Audit" sheet and change any "Criticality" value in column G (dropdown)',
            '2. Return here to see grade auto-update in cell B18',
            '3. Optionally edit "Org Test Coverage (%)" in cell B16 to simulate scenarios (enter 80 or 80%)',
            '4. Grade also auto-updates on "Executive Summary" (cell B3) and the "Metadata Snapshot" grade section',
            '5. Review the "Why This Grade?" section to see the exact blocker for the current grade',
            '6. Issue counts on "Metadata Snapshot" also auto-update via formulas',
            '7. To waive a loop finding from grading, mark that loop row as "False Positive" in either Criticality or Fix Status',
            '',
            'Formula Reference:',
            '  B7-B10 = COUNTIF on Detailed Audit column G for each severity',
            '  B11    = SUM of B7:B10',
            '  B13-B14 = SOQL/DML loop detection/count, excluding rows marked False Positive in Criticality or Fix Status',
            '  B16    = Org coverage input % (manually editable)',
            '  B17    = Normalized org coverage % used by grading formulas',
            '  B18    = Grade calculation (nested IF waterfall logic)',
            '  B19    = Grade rationale explanation',
        ]
        
        for idx, text in enumerate(instructions):
            r = 41 + idx
            ws[f'A{r}'] = text
            ws[f'A{r}'].font = Font(size=10)
            ws.merge_cells(f'A{r}:D{r}')
        
        # Column widths
        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 56
    
    def _create_rules_reference_sheet(self, wb: Workbook):
        """Create Rules Reference Sheet - Comprehensive rules documentation"""
        ws = wb.create_sheet('Rules Reference')
        
        # Title
        ws['A1'] = 'Salesforce Code Audit - Rules Reference'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 25
        
        # Subtitle
        ws['A2'] = f'All {len(self.RULES_REFERENCE)} violation types with descriptions, severity, and recommendations'
        ws['A2'].font = Font(size=11, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:F2')
        ws.row_dimensions[2].height = 20
        
        # Headers
        headers = ['#', 'Rule Name', 'Source', 'Category', 'Description', 'Severity', 'Recommendation']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Data rows
        for idx, rule_data in enumerate(self.RULES_REFERENCE, 1):
            row_num = idx + 3
            
            # Row number
            cell = ws.cell(row=row_num, column=1, value=idx)
            cell.alignment = Alignment(horizontal='center', vertical='top')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Rule Name
            cell = ws.cell(row=row_num, column=2, value=rule_data['rule'])
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Source
            cell = ws.cell(row=row_num, column=3, value=self._get_rule_source(rule_data['rule']))
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if cell.value == 'Apex Guru':
                cell.fill = PatternFill(start_color='D9EAF7', end_color='D9EAF7', fill_type='solid')

            # Category
            cell = ws.cell(row=row_num, column=4, value=rule_data['category'])
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Description
            cell = ws.cell(row=row_num, column=5, value=rule_data['description'])
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Severity with color coding
            severity = rule_data['severity']
            cell = ws.cell(row=row_num, column=6, value=severity)
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            cell.font = Font(bold=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Color code by severity
            if 'Critical' in severity:
                cell.fill = PatternFill(start_color=self.colors['critical'], end_color=self.colors['critical'], fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
            elif 'High' in severity:
                cell.fill = PatternFill(start_color=self.colors['high'], end_color=self.colors['high'], fill_type='solid')
            elif 'Medium' in severity:
                cell.fill = PatternFill(start_color=self.colors['medium'], end_color=self.colors['medium'], fill_type='solid')
            elif 'Low' in severity:
                cell.fill = PatternFill(start_color=self.colors['low'], end_color=self.colors['low'], fill_type='solid')
            
            # Recommendation
            cell = ws.cell(row=row_num, column=7, value=rule_data['recommendation'])
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Set row height for readability
            ws.row_dimensions[row_num].height = 60
        
        # Column widths
        ws.column_dimensions['A'].width = 5   # #
        ws.column_dimensions['B'].width = 28  # Rule Name
        ws.column_dimensions['C'].width = 12  # Source
        ws.column_dimensions['D'].width = 15  # Category
        ws.column_dimensions['E'].width = 50  # Description
        ws.column_dimensions['F'].width = 20  # Severity
        ws.column_dimensions['G'].width = 60  # Recommendation
        
        # Freeze panes
        ws.freeze_panes = 'A4'


class MarkdownReportGenerator:
    """Generates markdown summary reports"""
    
    def generate_summary(
        self,
        org_name: str,
        grading_result: GradingResult,
        output_dir: str = './audit_reports'
    ) -> str:
        """
        Generate markdown summary
        
        Args:
            org_name: Name of the Salesforce org
            grading_result: Grading result
            output_dir: Output directory
        
        Returns:
            Path to generated markdown file
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'SF_Audit_{org_name.replace(" ", "_")}_{timestamp}_Summary.md'
        filepath = os.path.join(output_dir, filename)
        
        markdown = self._generate_markdown_content(org_name, grading_result)
        
        with open(filepath, 'w') as f:
            f.write(markdown)
        
        print(f"✅ Markdown summary generated: {filepath}")
        return filepath
    
    def _generate_markdown_content(
        self,
        org_name: str,
        grading: GradingResult
    ) -> str:
        """Generate markdown content"""
        
        grade_emoji = {
            Grade.EXCELLENT: '🟢',
            Grade.VERY_GOOD: '🔵',
            Grade.GOOD: '🟡',
            Grade.AVERAGE: '🟠',
            Grade.BELOW_AVERAGE: '🔴',
            Grade.POOR: '🚨'
        }
        
        md = f"""# Salesforce Code Audit Summary

**Organization:** {org_name}  
**Audit Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

---

## Overall Code Health Grade

# {grade_emoji.get(grading.grade, '')} {grading.grade.value}

**Rationale:** {grading.grade_rationale}

---

## Executive Summary

{grading.executive_summary}

---

## Issue Breakdown

| Criticality | Count |
|-------------|-------|
| 🔴 Critical | {grading.issue_counts.critical} |
| 🟠 High     | {grading.issue_counts.high} |
| 🟡 Medium   | {grading.issue_counts.medium} |
| 🔵 Low      | {grading.issue_counts.low} |
| **Total**   | **{grading.issue_counts.total()}** |

---

## Test Coverage

- **Org Coverage:** {grading.coverage_stats.org_coverage:.1f}%
- **Classes Below 90%:** {grading.coverage_stats.classes_below_90}
- **Classes Below 75%:** {grading.coverage_stats.classes_below_75}
- **Average Class Coverage:** {grading.coverage_stats.average_class_coverage:.1f}%

---

## Top Priority Fixes

"""
        
        for i, fix in enumerate(grading.top_priority_fixes[:3], 1):
            md += f"""### {i}. {fix.get('rule', 'N/A')}

- **File:** `{fix.get('file', 'N/A')}`
- **Line:** {fix.get('line', 'N/A')}
- **Criticality:** {fix.get('criticality', 'N/A')}
- **Recommendation:** {fix.get('recommendation', 'N/A')}

"""
        
        md += """---

## Next Steps

"""
        
        if grading.grade == Grade.POOR:
            md += """1. **🚨 URGENT:** Halt new feature development until critical defects are resolved
2. **🚨 URGENT:** Address all SOQL/DML in loops immediately (production risk)
3. **🚨 URGENT:** Fix all Critical security vulnerabilities
4. **Immediate:** Increase test coverage to minimum 50%, target 75%
5. **Short-term:** Conduct comprehensive code review and establish refactoring plan
6. **Medium-term:** Implement code quality gates and mandatory peer reviews
7. **Long-term:** Establish technical debt reduction program
"""
        elif grading.grade in [Grade.BELOW_AVERAGE, Grade.AVERAGE]:
            md += """1. **Immediate:** Address all Critical issues, especially SOQL/DML in loops
2. **Short-term:** Improve test coverage to at least 75% for all classes
3. **Medium-term:** Resolve High priority issues
4. **Long-term:** Establish coding standards and peer review process
"""
        elif grading.grade == Grade.GOOD:
            md += """1. Address remaining High priority issues
2. Improve test coverage to 90% minimum
3. Review and fix Medium priority issues
4. Implement continuous code quality monitoring
"""
        else:
            md += """1. Continue maintaining high code quality standards
2. Address any remaining Medium priority issues
3. Implement automated quality gates in CI/CD pipeline
4. Share best practices with team
"""
        
        # Add Rules Reference Table
        md += "\n\n---\n\n"
        md += "## 📚 Rules Reference\n\n"
        md += f"*Comprehensive guide to all {len(ExcelReportGenerator.RULES_REFERENCE)} violation types*\n\n"
        md += "| # | Rule Name | Source | Category | Description | Severity | Recommendation |\n"
        md += "|---|-----------|--------|----------|-------------|----------|----------------|\n"
        
        for idx, rule in enumerate(ExcelReportGenerator.RULES_REFERENCE, 1):
            # Escape pipe characters in content
            rule_name = rule['rule'].replace('|', '\\|')
            source = ExcelReportGenerator._get_rule_source(rule['rule']).replace('|', '\\|')
            category = rule['category'].replace('|', '\\|')
            description = rule['description'].replace('|', '\\|')
            severity = rule['severity'].replace('|', '\\|')
            recommendation = rule['recommendation'].replace('|', '\\|')
            
            # Add severity emoji
            severity_emoji = ''
            if 'Critical' in severity:
                severity_emoji = '🔴'
            elif 'High' in severity:
                severity_emoji = '🟠'
            elif 'Medium' in severity:
                severity_emoji = '🟡'
            elif 'Low' in severity:
                severity_emoji = '🔵'
            
            md += f"| {idx} | **{rule_name}** | {source} | {category} | {description} | {severity_emoji} {severity} | {recommendation} |\n"
        
        md += "\n\n---\n\n"
        md += f"*Generated by Salesforce Code Audit Tool v{TOOL_VERSION} | {datetime.now().strftime('%B %d, %Y')}*\n"
        
        return md


class PDFReportGenerator:
    """Generates PDF executive summary reports with Salesforce branding"""
    
    def __init__(self):
        """Initialize PDF report generator"""
        self.salesforce_blue = (0/255, 161/255, 224/255)  # #00A1E0
        self.critical_red = (229/255, 43/255, 26/255)     # #E52B1A
        self.high_orange = (255/255, 154/255, 60/255)     # #FF9A3C
        self.medium_yellow = (255/255, 183/255, 93/255)   # #FFB75D
        self.low_blue = (21/255, 137/255, 238/255)        # #1589EE

    @staticmethod
    def _clean_pdf_text(text: Any) -> str:
        """Convert mixed plain/html-ish content into clean readable text."""
        if text is None:
            return ""

        cleaned = str(text)
        cleaned = cleaned.replace("\r\n", "\n").replace("\r", "\n")
        cleaned = re.sub(r'<br\s*/?>', '\n', cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r'</?(b|strong|i|em|u)>', '', cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r'<[^>]+>', '', cleaned)
        cleaned = re.sub(r'[ \t]+', ' ', cleaned)
        cleaned = re.sub(r' *\n *', '\n', cleaned)
        cleaned = re.sub(r'\n{3,}', '\n\n', cleaned)
        return cleaned.strip()

    def _paragraph_from_text(self, text: Any, style):
        """Build a ReportLab paragraph from plain text while preserving line breaks."""
        from reportlab.platypus import Paragraph

        cleaned = self._clean_pdf_text(text)
        safe_text = escape(cleaned).replace('\n', '<br/>')
        return Paragraph(safe_text or "N/A", style)

    def _build_paragraph_block(self, story: List[Any], text: Any, style, spacer_height: float):
        """Split a text block into readable paragraphs."""
        from reportlab.platypus import Spacer

        cleaned = self._clean_pdf_text(text)
        parts = [part.strip() for part in re.split(r'\n\s*\n', cleaned) if part.strip()]
        if not parts:
            parts = ["No summary available."]

        for part in parts:
            story.append(self._paragraph_from_text(part, style))
            story.append(Spacer(1, spacer_height))

    def _summarize_recommendation(self, text: Any, limit: int = 180) -> str:
        """Condense long recommendations for PDF cards."""
        cleaned = self._clean_pdf_text(text)
        if not cleaned:
            return "Review and remediate this issue."

        first_line = cleaned.split('\n', 1)[0].strip()
        first_sentence = re.split(r'(?<=[.!?])\s+', first_line, maxsplit=1)[0].strip()
        summary = first_sentence or first_line
        if len(summary) > limit:
            summary = summary[: limit - 3].rstrip() + "..."
        return summary

    def _build_action_plan_sections(self, grading_result) -> List[tuple]:
        """Create a concise action plan tailored to the current grading result."""
        issues = grading_result.issue_counts
        coverage = grading_result.coverage_stats.org_coverage
        top_fixes = grading_result.top_priority_fixes[:3] if grading_result.top_priority_fixes else []
        top_rules = [self._clean_pdf_text(fix.get('rule', 'Unknown')) for fix in top_fixes]

        immediate = []
        near_term = []
        strategic = []

        if grading_result.grade.value == 'Poor':
            immediate.append("Stabilize production risk items before further feature work.")
        if coverage < 75:
            immediate.append(f"Raise org test coverage from {coverage:.1f}% to at least 75% on critical execution paths.")
        if issues.critical > 0:
            immediate.append(f"Resolve the {issues.critical} critical findings first, starting with the highest-volume rules.")
        if any('Loop' in rule for rule in top_rules):
            immediate.append("Bulkify loop-driven database operations and remove SOQL/DML execution from loops.")

        if issues.high > 0:
            near_term.append(f"Reduce the {issues.high} high-severity issues to keep remediation from stalling delivery.")
        near_term.append("Harden CRUD/FLS and sharing enforcement on externally reachable classes and controllers.")
        near_term.append("Refactor repeated anti-patterns into reusable guardrails and helper utilities.")

        strategic.append("Add automated audit review to release readiness so new critical findings are blocked earlier.")
        strategic.append("Track the top recurring rules by owner/team and burn them down systematically.")
        if top_rules:
            strategic.append(f"Prioritize recurring rule families such as {', '.join(top_rules[:2])}.")

        return [
            ("Immediate", immediate[:4]),
            ("Next Sprint", near_term[:4]),
            ("Sustained Improvements", strategic[:4]),
        ]

    @staticmethod
    def _get_next_grade_guidance(current_grade: str) -> Dict[str, str]:
        """Return the next grade target and the main blocker guidance."""
        guidance = {
            'Poor': {
                'next_grade': 'Below Average',
                'blocker': 'Fix the Poor-grade blocker first: remove SOQL/DML loop issues, raise coverage above 50%, and bring Critical issues below 10.',
            },
            'Below Average': {
                'next_grade': 'Average',
                'blocker': 'Raise coverage to at least 70%, reduce Critical issues to 5 or fewer, and keep High issues below 50.',
            },
            'Average': {
                'next_grade': 'Good',
                'blocker': 'Raise coverage to at least 75%, reduce Critical issues to 3 or fewer, and keep High issues below 15.',
            },
            'Good': {
                'next_grade': 'Very Good',
                'blocker': 'Raise coverage to at least 85%, reduce Critical issues to 0, keep High issues below 10, and keep Medium issues below 30.',
            },
            'Very Good': {
                'next_grade': 'Excellent',
                'blocker': 'Raise coverage to at least 90%, reduce High issues to 0, and keep Medium issues below 15.',
            },
            'Excellent': {
                'next_grade': 'Excellent',
                'blocker': 'Already at the highest grade band.',
            },
        }
        return guidance.get(current_grade, {
            'next_grade': 'Unknown',
            'blocker': 'Review issue counts and coverage against the grading thresholds.',
        })

    @staticmethod
    def _get_grading_reference_rows() -> List[List[str]]:
        """Compact grading band reference for the PDF."""
        return [
            ['Excellent', '90%+', '0 Critical, 0 High, <15 Medium'],
            ['Very Good', '85%+', '0 Critical, <10 High, <30 Medium'],
            ['Good', '75%+', '<=3 Critical, <15 High'],
            ['Average', '70%+', '<=5 Critical, <50 High'],
            ['Below Average', '50%+', '<=10 Critical'],
            ['Poor', '<50% or loop blocker', '10+ Critical or SOQL/DML loops'],
        ]
    
    def generate_executive_summary(
        self,
        org_name: str,
        grading_result: GradingResult,
        metadata_snapshot: Dict,
        output_dir: str = './audit_reports'
    ) -> str:
        """
        Generate PDF executive summary with Salesforce branding
        
        Args:
            org_name: Name of the Salesforce org
            grading_result: Grading result from GradingEngine
            metadata_snapshot: Org metadata stats
            output_dir: Output directory
        
        Returns:
            Path to generated PDF file
        """
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'SF_Audit_{org_name.replace(" ", "_")}_{timestamp}_Executive_Summary.pdf'
            os.makedirs(output_dir, exist_ok=True)
            filepath = os.path.join(output_dir, filename)
            
            # Create PDF document
            doc = SimpleDocTemplate(
                filepath,
                pagesize=letter,
                rightMargin=0.75*inch,
                leftMargin=0.75*inch,
                topMargin=0.75*inch,
                bottomMargin=0.75*inch
            )
            
            # Container for PDF elements
            story = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#00A1E0'),
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )

            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['BodyText'],
                fontSize=13,
                textColor=colors.grey,
                alignment=TA_CENTER,
                leading=16
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#00A1E0'),
                spaceAfter=12,
                spaceBefore=12,
                fontName='Helvetica-Bold'
            )

            body_style = ParagraphStyle(
                'PdfBody',
                parent=styles['BodyText'],
                fontSize=10.5,
                leading=14,
                spaceAfter=6
            )

            note_style = ParagraphStyle(
                'PdfNote',
                parent=styles['BodyText'],
                fontSize=9,
                leading=12,
                textColor=colors.grey
            )

            callout_style = ParagraphStyle(
                'PdfCallout',
                parent=styles['BodyText'],
                fontSize=10.5,
                leading=14,
                borderPadding=10,
                backColor=colors.HexColor('#F4F9FC')
            )
            
            # Page 1: Cover Page
            story.append(Spacer(1, 1.1*inch))
            
            logo_text = Paragraph(
                "<b>SALESFORCE</b>",
                ParagraphStyle(
                    'Logo',
                    fontSize=32,
                    textColor=colors.HexColor('#00A1E0'),
                    alignment=TA_CENTER,
                    fontName='Helvetica-Bold'
                )
            )
            story.append(logo_text)
            story.append(Spacer(1, 0.35*inch))
            
            story.append(Paragraph("SALESFORCE CODE AUDIT REPORT", title_style))
            story.append(Paragraph("Executive summary for leadership review", subtitle_style))
            story.append(Spacer(1, 0.45*inch))
            
            total_files = metadata_snapshot.get('apex_classes', 0) + metadata_snapshot.get('triggers', 0)
            coverage_text = f"{grading_result.coverage_stats.org_coverage:.1f}%"
            audit_timestamp = metadata_snapshot.get('audit_timestamp', datetime.now().isoformat())
            org_details = [
                [self._paragraph_from_text("Audit Timestamp", styles['Heading4']), self._paragraph_from_text(audit_timestamp, body_style)],
                [self._paragraph_from_text("Org Name", styles['Heading4']), self._paragraph_from_text(metadata_snapshot.get('org_name', org_name), body_style)],
                [self._paragraph_from_text("Org Type", styles['Heading4']), self._paragraph_from_text(metadata_snapshot.get('org_type', 'N/A'), body_style)],
                [self._paragraph_from_text("API Version", styles['Heading4']), self._paragraph_from_text(metadata_snapshot.get('api_version', 'N/A'), body_style)],
                [self._paragraph_from_text("Org Domain Name", styles['Heading4']), self._paragraph_from_text(metadata_snapshot.get('org_domain_name', 'N/A'), body_style)],
                [self._paragraph_from_text("Files Analyzed", styles['Heading4']), self._paragraph_from_text(str(total_files), body_style)],
                [self._paragraph_from_text("Org Coverage", styles['Heading4']), self._paragraph_from_text(coverage_text, body_style)],
                [self._paragraph_from_text("Tool Version", styles['Heading4']), self._paragraph_from_text(f"v{TOOL_VERSION}", body_style)],
            ]
            
            org_table = Table(org_details, colWidths=[1.8*inch, 4.1*inch])
            org_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F4F9FC')),
                ('BOX', (0, 0), (-1, -1), 0.75, colors.HexColor('#D5EAF4')),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D5EAF4')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(org_table)
            
            grade_colors = {
                'Excellent': colors.HexColor('#90EE90'),
                'Very Good': colors.HexColor('#C6E0B4'),
                'Good': colors.HexColor('#FFB75D'),
                'Average': colors.HexColor('#FF9A3C'),
                'Below Average': colors.HexColor('#E52B1A'),
                'Poor': colors.HexColor('#8B0000')
            }
            
            grade_color = grade_colors.get(grading_result.grade.value, colors.grey)
            
            story.append(Spacer(1, 0.35*inch))
            grade_box = Table(
                [[
                    self._paragraph_from_text(
                        f"{grading_result.grade.value}\n{grading_result.grade_rationale}",
                        ParagraphStyle(
                            'GradeBoxText',
                            parent=body_style,
                            alignment=TA_CENTER,
                            fontName='Helvetica-Bold',
                            textColor=colors.white if grading_result.grade.value in ['Below Average', 'Poor'] else colors.black,
                            leading=15
                        )
                    )
                ]],
                colWidths=[5.5*inch],
                rowHeights=[0.95*inch]
            )
            grade_box.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BACKGROUND', (0, 0), (-1, -1), grade_color),
                ('BOX', (0, 0), (-1, -1), 2, colors.black),
                ('LEFTPADDING', (0, 0), (-1, -1), 12),
                ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ]))
            story.append(grade_box)
            story.append(Spacer(1, 0.2*inch))
            story.append(self._paragraph_from_text(
                "This PDF focuses on the overall health signal, the highest-risk findings, and the recommended remediation sequence.",
                note_style
            ))
            
            story.append(PageBreak())
            
            story.append(Paragraph("EXECUTIVE SUMMARY", heading_style))
            story.append(Spacer(1, 0.2*inch))

            self._build_paragraph_block(story, grading_result.executive_summary, body_style, 0.08 * inch)

            story.append(Paragraph("KEY METRICS", heading_style))
            total_issues = grading_result.issue_counts.total()
            metrics_data = [
                [self._paragraph_from_text("Metric", styles['Heading4']), self._paragraph_from_text("Value", styles['Heading4'])],
                ['Files Analyzed', str(total_files)],
                ['Total Issues', str(total_issues)],
                ['Critical Issues', f"{grading_result.issue_counts.critical} ({grading_result.issue_counts.critical/max(total_issues, 1)*100:.1f}%)"],
                ['High Issues', f"{grading_result.issue_counts.high} ({grading_result.issue_counts.high/max(total_issues, 1)*100:.1f}%)"],
                ['Test Coverage', f"{grading_result.coverage_stats.org_coverage:.1f}%"],
                ['Classes Below 75%', str(grading_result.coverage_stats.classes_below_75)]
            ]

            metrics_data = [
                row if index == 0 else [self._paragraph_from_text(row[0], body_style), self._paragraph_from_text(row[1], body_style)]
                for index, row in enumerate(metrics_data)
            ]
            metrics_table = Table(metrics_data, colWidths=[3*inch, 3*inch])
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00A1E0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9F4EA')),
                ('GRID', (0, 0), (-1, -1), 0.75, colors.HexColor('#D9D9D9')),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ]))
            story.append(metrics_table)
            story.append(Spacer(1, 0.18 * inch))
            story.append(self._paragraph_from_text(grading_result.grade_rationale, callout_style))
            story.append(Spacer(1, 0.15 * inch))

            next_grade_guidance = self._get_next_grade_guidance(grading_result.grade.value)
            story.append(Paragraph("WHY THIS GRADE?", heading_style))
            why_grade_rows = [
                [self._paragraph_from_text("Current Grade", styles['Heading4']), self._paragraph_from_text(grading_result.grade.value, body_style)],
                [self._paragraph_from_text("Why This Grade?", styles['Heading4']), self._paragraph_from_text(grading_result.grade_rationale, body_style)],
                [self._paragraph_from_text("Next Better Grade", styles['Heading4']), self._paragraph_from_text(next_grade_guidance['next_grade'], body_style)],
                [self._paragraph_from_text("What Blocks It", styles['Heading4']), self._paragraph_from_text(next_grade_guidance['blocker'], body_style)],
            ]
            why_grade_table = Table(why_grade_rows, colWidths=[1.8*inch, 4.2*inch])
            why_grade_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F4F9FC')),
                ('BOX', (0, 0), (-1, -1), 0.75, colors.HexColor('#D5EAF4')),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D5EAF4')),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(why_grade_table)

            story.append(PageBreak())
            
            story.append(Paragraph("ISSUE BREAKDOWN", heading_style))
            story.append(Spacer(1, 0.2*inch))
            
            issue_data = [
                [self._paragraph_from_text("Severity", styles['Heading4']),
                 self._paragraph_from_text("Count", styles['Heading4']),
                 self._paragraph_from_text("Share", styles['Heading4'])],
                ['Critical', str(grading_result.issue_counts.critical), f"{grading_result.issue_counts.critical/max(total_issues, 1)*100:.1f}%"],
                ['High', str(grading_result.issue_counts.high), f"{grading_result.issue_counts.high/max(total_issues, 1)*100:.1f}%"],
                ['Medium', str(grading_result.issue_counts.medium), f"{grading_result.issue_counts.medium/max(total_issues, 1)*100:.1f}%"],
                ['Low', str(grading_result.issue_counts.low), f"{grading_result.issue_counts.low/max(total_issues, 1)*100:.1f}%"],
            ]

            issue_data = [
                row if index == 0 else [self._paragraph_from_text(cell, body_style) for cell in row]
                for index, row in enumerate(issue_data)
            ]
            issue_table = Table(issue_data, colWidths=[2*inch, 2*inch, 2*inch])
            issue_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00A1E0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.75, colors.HexColor('#D9D9D9')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F6F6F6')]),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(issue_table)
            story.append(Spacer(1, 0.3*inch))
            
            story.append(Paragraph("TOP PRIORITY FIXES", heading_style))
            top_fixes = grading_result.top_priority_fixes[:5] if grading_result.top_priority_fixes else []

            for i, fix in enumerate(top_fixes, 1):
                rule_name = self._clean_pdf_text(fix.get('rule', 'N/A'))
                rule_name = re.sub(r'\s*\(\d+\s+instances\)\s*$', '', rule_name)
                count = fix.get('count', 0)
                severity = self._clean_pdf_text(fix.get('criticality', 'N/A'))
                file_examples = self._clean_pdf_text(fix.get('file', ''))
                recommendation = self._summarize_recommendation(fix.get('recommendation', ''))

                card_rows = [
                    [self._paragraph_from_text(f"{i}. {rule_name}", ParagraphStyle('FixTitle', parent=body_style, fontName='Helvetica-Bold', fontSize=11, leading=14)),
                     self._paragraph_from_text(f"{count} instance(s) | {severity}", ParagraphStyle('FixMeta', parent=note_style, alignment=TA_RIGHT))],
                    [self._paragraph_from_text(recommendation, body_style), self._paragraph_from_text(f"Examples: {file_examples}" if file_examples else "", note_style)],
                ]
                card = Table(card_rows, colWidths=[4.15*inch, 1.85*inch])
                card.setStyle(TableStyle([
                    ('BOX', (0, 0), (-1, -1), 0.75, colors.HexColor('#D5EAF4')),
                    ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D5EAF4')),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F4F9FC')),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ]))
                story.append(card)
                story.append(Spacer(1, 0.12 * inch))

            story.append(PageBreak())
            
            story.append(Paragraph("RECOMMENDED ACTION PLAN", heading_style))
            story.append(Spacer(1, 0.2*inch))

            action_sections = self._build_action_plan_sections(grading_result)
            for title, items in action_sections:
                story.append(self._paragraph_from_text(title, ParagraphStyle('ActionTitle', parent=styles['Heading3'], textColor=colors.HexColor('#00A1E0'))))
                for item in items:
                    story.append(self._paragraph_from_text(f"- {item}", body_style))
                story.append(Spacer(1, 0.08 * inch))

            story.append(Paragraph("SCOPE SNAPSHOT", heading_style))
            scope_rows = [
                [self._paragraph_from_text("Apex Classes", body_style), self._paragraph_from_text(str(metadata_snapshot.get('apex_classes', 0)), body_style)],
                [self._paragraph_from_text("Triggers", body_style), self._paragraph_from_text(str(metadata_snapshot.get('triggers', 0)), body_style)],
                [self._paragraph_from_text("Flows", body_style), self._paragraph_from_text(str(metadata_snapshot.get('flows', 0)), body_style)],
                [self._paragraph_from_text("LWC Bundles", body_style), self._paragraph_from_text(str(metadata_snapshot.get('lwc_components', 0)), body_style)],
            ]
            scope_table = Table(scope_rows, colWidths=[3.5*inch, 2.5*inch])
            scope_table.setStyle(TableStyle([
                ('BOX', (0, 0), (-1, -1), 0.75, colors.HexColor('#D9D9D9')),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
                ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F6F6F6')]),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(scope_table)
            story.append(Spacer(1, 0.18 * inch))

            story.append(Paragraph("GRADING REFERENCE", heading_style))
            grading_reference_rows = [
                [self._paragraph_from_text("Grade", styles['Heading4']),
                 self._paragraph_from_text("Coverage", styles['Heading4']),
                 self._paragraph_from_text("Reference", styles['Heading4'])]
            ]
            for row in self._get_grading_reference_rows():
                grading_reference_rows.append([self._paragraph_from_text(cell, body_style) for cell in row])

            grading_reference_table = Table(grading_reference_rows, colWidths=[1.5*inch, 1.35*inch, 3.15*inch])
            grading_reference_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00A1E0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('GRID', (0, 0), (-1, -1), 0.75, colors.HexColor('#D9D9D9')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F6F6F6')]),
                ('LEFTPADDING', (0, 0), (-1, -1), 7),
                ('RIGHTPADDING', (0, 0), (-1, -1), 7),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(grading_reference_table)

            story.append(Spacer(1, 0.3*inch))
            footer_text = f"<i>Generated by Salesforce Code Audit Tool v{TOOL_VERSION} | {datetime.now().strftime('%B %d, %Y')}</i>"
            story.append(Paragraph(footer_text, ParagraphStyle(
                'Footer',
                fontSize=9,
                textColor=colors.grey,
                alignment=TA_CENTER
            )))
            
            # Build PDF
            doc.build(story)
            
            print(f"✅ PDF Executive Summary generated: {filepath}")
            return filepath
            
        except ImportError:
            print("⚠️  ReportLab library not installed. Skipping PDF generation.")
            print("   Install with: pip install reportlab")
            return None
        except Exception as e:
            print(f"⚠️  PDF generation failed: {str(e)}")
            return None


if __name__ == "__main__":
    print("Report Generator Module - Use via main audit script")

